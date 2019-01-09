import names
from openpyxl import Workbook
from msxlsxparsing.Node import *
import random


def createExcel():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 'First'
    ws['B1'] = 'Last'
    ws['C1'] = 'Street'
    ws['D1'] = 'Phone'
    ws['E1'] = 'Email'

    node = Node('John', 'Doe', '123 Main St', '555-333-3333', 'gtohill@localhost.com')

    phone_numbers = list()
    for x in range(0,10):
        phone_numbers.append(x)

    street_suffix = ['St', 'Ave', 'Cres', 'Rd', 'Blvd', 'Line']
    email_host = ['gmail.com', 'yahoo.com', 'bell.net', 'hotmail.com', 'sprint.com']

    # create random person, street, phone, and email
    node_list = list()
    r = 2
    for x in range(100):
        first = names.get_first_name()
        last = names.get_last_name()
        street = str(random.randint(1,9)+random.randint(0,9)+random.randint(0,9))+' '+names.get_last_name()+' '+random.choice(street_suffix)
        email = first+'@'+random.choice(email_host)

        phone = ''
        for j in range(0,12):
            if j == 3 or j == 7:
                phone += '-'
            else:
                if j == 0:
                    phone += str(random.randint(2, 9))
                else:
                    phone += str(random.randint(0,9))

        ws['A'+str(r)] = first
        ws['B'+str(r)] = last
        ws['C'+str(r)] = street
        ws['D'+str(r)] = phone
        ws['E'+str(r)] = email

        # increment row counter(r)
        r += 1

        node = Node(first, last, street, phone, email)
        node_list.append(node)

    # Save the file
    wb.save("sample.xlsx")