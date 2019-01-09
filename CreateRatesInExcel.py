import names
from openpyxl import Workbook
from msxlsxparsing.Node import *
import random


def createExcelSpreadSheet():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    '''
    ws['A1'] = 'Company'
    ws['B1'] = '1 year'
    ws['C1'] = '2 year'
    ws['D1'] = '3 year'
    ws['E1'] = '4 year'
    ws['F1'] = '5 year'
    '''

    blist = open('banklist', 'r').read()
    bank_list = blist.split(',')
    banks = list()
    for x in range(len(bank_list)):
        if x%2 == 1:
            banks.append(bank_list[x])

    # base rates. use to calculate annual rates for each bank. a random factor will be added or subtracted.
    oneYearRate = .015
    twoYearRate = .0175
    threeYearRate = .02
    fourYearRate = .0225
    fiveYearRate = .025

    r = 1
    for bank in banks:
        one = round(oneYearRate+random.uniform(-.005, .005), 4)
        two = round(twoYearRate+random.uniform(-.005, .005), 4)
        three = round(threeYearRate + random.uniform(-.005, .005), 4)
        four = round(fourYearRate + random.uniform(-.005, .005), 4)
        five = round(fiveYearRate + random.uniform(-.005, .005), 4)

        ws['A'+str(r)] = bank
        ws['B'+str(r)] = one
        ws['C'+str(r)] = two
        ws['D'+str(r)] = three
        ws['E'+str(r)] = four
        ws['F'+str(r)] = five

        # increment row counter(r)
        r += 1
    
    # Save the file
    wb.save("rates.xlsx")


if __name__ == '__main__':

    createExcelSpreadSheet()
